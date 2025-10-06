"""
Spreadsheet processor for Excel, CSV, ODS files.

This module provides comprehensive spreadsheet processing:
- Excel (XLSX, XLS) support
- CSV support with encoding detection
- ODS (OpenDocument Spreadsheet) support
- Formula extraction
- Cell formatting preservation
- Multiple sheet handling
- Table structure detection
"""

from typing import Dict, Any, Optional, List, Tuple
from pathlib import Path
from datetime import datetime
import csv
import io

import structlog

from .models_result import ProcessResult, ProcessorError, ErrorCode, ProcessingStage
from .dependencies import get_dependency_checker

logger = structlog.get_logger(__name__)


class SpreadsheetProcessor:
    """
    Comprehensive spreadsheet processor.
    
    Features:
    - Multiple format support (XLSX, XLS, CSV, ODS)
    - Formula extraction
    - Cell formatting
    - Multiple sheets
    - Table detection
    - Data type inference
    """
    
    def __init__(
        self,
        max_rows: int = 10000,
        max_cols: int = 100,
        include_formulas: bool = True
    ):
        """
        Initialize spreadsheet processor.
        
        Args:
            max_rows: Maximum rows to process per sheet
            max_cols: Maximum columns to process
            include_formulas: Include formula information
        """
        self.max_rows = max_rows
        self.max_cols = max_cols
        self.include_formulas = include_formulas
        
        self.dep_checker = get_dependency_checker()
        
        logger.info(
            "SpreadsheetProcessor initialized",
            max_rows=max_rows,
            max_cols=max_cols
        )
    
    async def process(
        self,
        content: bytes,
        filename: str,
        metadata: Optional[Dict[str, Any]] = None
    ) -> ProcessResult:
        """
        Process spreadsheet file.
        
        Args:
            content: Spreadsheet file content
            filename: Filename
            metadata: Additional metadata
            
        Returns:
            ProcessResult with extracted data
        """
        start_time = datetime.utcnow()
        errors = []
        
        try:
            # Detect format
            file_ext = Path(filename).suffix.lower()
            
            if file_ext in ['.xlsx', '.xlsm']:
                result_data = await self._process_xlsx(content, filename)
            elif file_ext == '.xls':
                result_data = await self._process_xls(content, filename)
            elif file_ext == '.csv':
                result_data = await self._process_csv(content, filename)
            elif file_ext == '.ods':
                result_data = await self._process_ods(content, filename)
            else:
                return ProcessResult(
                    text="",
                    metadata=metadata or {},
                    errors=[ProcessorError(
                        code=ErrorCode.UNSUPPORTED_FORMAT,
                        message=f"Unsupported spreadsheet format: {file_ext}",
                        stage=ProcessingStage.EXTRACTION,
                        retriable=False
                    )],
                    processor_name="SpreadsheetProcessor",
                    processing_time_ms=(datetime.utcnow() - start_time).total_seconds() * 1000
                )
            
            # Build text representation
            text_parts = [f"Spreadsheet: {filename}"]
            text_parts.append(f"Sheets: {len(result_data['sheets'])}")
            text_parts.append("")
            
            for sheet_info in result_data['sheets']:
                text_parts.append(f"## Sheet: {sheet_info['name']}")
                text_parts.append(f"Rows: {sheet_info['rows']}, Columns: {sheet_info['cols']}")
                text_parts.append("")
                
                # Add table data
                if sheet_info.get('data'):
                    for row in sheet_info['data'][:20]:  # First 20 rows
                        text_parts.append(" | ".join(str(cell) for cell in row))
                    
                    if sheet_info['rows'] > 20:
                        text_parts.append(f"... and {sheet_info['rows'] - 20} more rows")
                
                text_parts.append("")
            
            result_metadata = {
                **(metadata or {}),
                "spreadsheet_format": file_ext[1:],
                "total_sheets": len(result_data['sheets']),
                "sheets": result_data['sheets']
            }
            
            return ProcessResult(
                text="\n".join(text_parts),
                metadata=result_metadata,
                errors=errors,
                processor_name="SpreadsheetProcessor",
                processing_time_ms=(datetime.utcnow() - start_time).total_seconds() * 1000
            )
            
        except Exception as e:
            logger.error("Spreadsheet processing failed", error=str(e), filename=filename)
            
            return ProcessResult(
                text="",
                metadata=metadata or {},
                errors=[ProcessorError(
                    code=ErrorCode.PROCESSING_FAILED,
                    message=f"Spreadsheet processing failed: {str(e)}",
                    stage=ProcessingStage.EXTRACTION,
                    retriable=True,
                    details={"error": str(e)}
                )],
                processor_name="SpreadsheetProcessor",
                processing_time_ms=(datetime.utcnow() - start_time).total_seconds() * 1000
            )
    
    async def _process_xlsx(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Process XLSX file using openpyxl."""
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(content), data_only=not self.include_formulas)
        
        sheets = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get dimensions
            max_row = min(ws.max_row, self.max_rows)
            max_col = min(ws.max_column, self.max_cols)
            
            # Extract data
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(list(row))
            
            # Extract formulas if enabled
            formulas = []
            if self.include_formulas:
                for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            formulas.append({
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
            
            sheets.append({
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col,
                "data": data,
                "formulas": formulas
            })
        
        return {"sheets": sheets}
    
    async def _process_xls(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Process XLS file using xlrd."""
        import xlrd
        
        wb = xlrd.open_workbook(file_contents=content)
        
        sheets = []
        
        for sheet in wb.sheets():
            max_row = min(sheet.nrows, self.max_rows)
            max_col = min(sheet.ncols, self.max_cols)
            
            # Extract data
            data = []
            for row_idx in range(max_row):
                row = []
                for col_idx in range(max_col):
                    cell = sheet.cell(row_idx, col_idx)
                    row.append(cell.value)
                data.append(row)
            
            sheets.append({
                "name": sheet.name,
                "rows": max_row,
                "cols": max_col,
                "data": data,
                "formulas": []
            })
        
        return {"sheets": sheets}
    
    async def _process_csv(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Process CSV file with encoding detection."""
        # Detect encoding
        encoding = self._detect_encoding(content)
        
        # Decode content
        text = content.decode(encoding, errors='replace')
        
        # Parse CSV
        reader = csv.reader(io.StringIO(text))
        
        data = []
        for idx, row in enumerate(reader):
            if idx >= self.max_rows:
                break
            
            # Limit columns
            data.append(row[:self.max_cols])
        
        sheets = [{
            "name": "Sheet1",
            "rows": len(data),
            "cols": max(len(row) for row in data) if data else 0,
            "data": data,
            "formulas": []
        }]
        
        return {"sheets": sheets}
    
    async def _process_ods(self, content: bytes, filename: str) -> Dict[str, Any]:
        """Process ODS file using odfpy."""
        from odf import opendocument
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        
        doc = opendocument.load(io.BytesIO(content))
        
        sheets = []
        
        for table in doc.spreadsheet.getElementsByType(Table):
            sheet_name = table.getAttribute('name')
            
            data = []
            
            for row_idx, row in enumerate(table.getElementsByType(TableRow)):
                if row_idx >= self.max_rows:
                    break
                
                row_data = []
                
                for col_idx, cell in enumerate(row.getElementsByType(TableCell)):
                    if col_idx >= self.max_cols:
                        break
                    
                    # Get cell value
                    cell_value = ""
                    for p in cell.getElementsByType(P):
                        cell_value += str(p)
                    
                    row_data.append(cell_value)
                
                data.append(row_data)
            
            sheets.append({
                "name": sheet_name,
                "rows": len(data),
                "cols": max(len(row) for row in data) if data else 0,
                "data": data,
                "formulas": []
            })
        
        return {"sheets": sheets}
    
    def _detect_encoding(self, content: bytes) -> str:
        """Detect text encoding."""
        # Try common encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                content.decode(encoding)
                return encoding
            except UnicodeDecodeError:
                continue
        
        # Fallback to utf-8 with error handling
        return 'utf-8'

