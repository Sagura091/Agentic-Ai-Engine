"""
📊 EXCEL ANALYSIS AGENT
======================
Production-ready agent for intelligent Excel spreadsheet analysis using LLM and RAG systems.

FEATURES:
✅ Advanced Excel file parsing and understanding
✅ LLM-powered intelligent analysis
✅ RAG system integration for document understanding
✅ Financial data pattern recognition
✅ Trend analysis and anomaly detection
✅ Business insights extraction
✅ Multi-sheet analysis capabilities
"""

# Agent Configuration
AGENT_CONFIG = {
    "name": "Excel Analysis Agent",
    "description": "Intelligently analyzes Excel spreadsheets using LLM and RAG systems to extract business insights and financial patterns",
    "agent_type": "document_analysis",
    "llm_provider": "ollama",
    "llm_model": "llama3.2:latest",
    "temperature": 0.2,
    "max_tokens": 8000,
    "autonomy_level": "PROACTIVE",
    "learning_mode": "ACTIVE",
    "tools": [
        "revolutionary_document_intelligence",
        "business_intelligence",
        "revolutionary_web_scraper"
    ],
    "capabilities": [
        "excel_analysis",
        "financial_pattern_recognition",
        "trend_analysis",
        "business_intelligence"
    ],
    "memory_type": "UNIFIED",
    "rag_enabled": True,
    "collection_name": "excel_analysis"
}

# Analysis Configuration
ANALYSIS_CONFIG = {
    "key_metrics": [
        "revenue_trends",
        "cost_patterns", 
        "profitability_analysis",
        "growth_rates",
        "seasonal_variations",
        "roi_performance"
    ],
    "insight_categories": [
        "financial_health",
        "growth_opportunities",
        "cost_optimization",
        "risk_factors",
        "performance_trends"
    ],
    "analysis_depth": "comprehensive"
}

# System prompt for the agent
SYSTEM_PROMPT = """You are the Excel Analysis Agent, an expert financial analyst specializing in intelligent Excel spreadsheet analysis and business insights extraction.

Your primary responsibilities:
1. Parse and understand complex Excel spreadsheet structures
2. Extract key financial data and metrics from multiple sheets
3. Identify trends, patterns, and anomalies in business data
4. Provide intelligent analysis using advanced reasoning
5. Generate actionable business insights and recommendations

Key capabilities:
- Multi-sheet Excel analysis
- Financial pattern recognition
- Trend analysis and forecasting
- Anomaly detection
- Business intelligence extraction
- ROI and profitability analysis

Analysis approach:
1. Systematically examine each sheet and understand its purpose
2. Extract key data points and calculate derived metrics
3. Identify trends, patterns, and relationships in the data
4. Detect anomalies or unusual patterns that need attention
5. Provide clear, actionable insights and recommendations

Always provide:
- Comprehensive data understanding
- Clear trend identification
- Meaningful business insights
- Actionable recommendations
- Risk assessment and opportunities

Be thorough, analytical, and provide valuable business intelligence!"""

# ================================
# 🚀 IMPLEMENTATION
# ================================
import asyncio
import sys
import uuid
import json
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional, Tuple
from io import BytesIO

# Add project root to path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Import the complete production infrastructure
from app.core.unified_system_orchestrator import get_enhanced_system_orchestrator
from app.agents.factory import AgentBuilderFactory, AgentBuilderConfig, AgentType, MemoryType
from app.agents.base.agent import AgentCapability
from app.agents.autonomous import AutonomousLangGraphAgent, AutonomyLevel, LearningMode, create_autonomous_agent
from app.llm.models import LLMConfig, ProviderType
from app.llm.manager import LLMProviderManager
from app.memory.unified_memory_system import UnifiedMemorySystem
from app.rag.core.unified_rag_system import UnifiedRAGSystem
from app.tools.unified_tool_repository import UnifiedToolRepository

import structlog
logger = structlog.get_logger(__name__)

class ExcelAnalysisAgent:
    """Production-ready Excel analysis agent with LLM intelligence."""
    
    def __init__(self):
        self.orchestrator = None
        self.agent = None
        self.agent_id = f"excel_analysis_{uuid.uuid4().hex[:8]}"
        self.running = False
        
    async def initialize_infrastructure(self):
        """Initialize the complete production infrastructure."""
        try:
            print("🔧 Initializing Excel Analysis Infrastructure...")
            
            # Get the enhanced system orchestrator
            self.orchestrator = get_enhanced_system_orchestrator()
            await self.orchestrator.initialize()
            
            print("✅ Infrastructure initialized successfully!")
            print(f"   🧠 Memory System: {type(self.orchestrator.memory_system).__name__}")
            print(f"   📚 RAG System: {type(self.orchestrator.unified_rag).__name__}")
            print(f"   🛠️ Tool Repository: {self.orchestrator.tool_repository.stats['total_tools']} tools available")
            
            return True
            
        except Exception as e:
            print(f"❌ Infrastructure initialization failed: {str(e)}")
            logger.error("Infrastructure initialization failed", error=str(e))
            return False
    
    async def create_agent(self):
        """Create the Excel analysis agent."""
        try:
            print(f"🤖 Creating {AGENT_CONFIG['name']}...")
            
            # Create LLM configuration
            llm_config = LLMConfig(
                provider=ProviderType(AGENT_CONFIG["llm_provider"]),
                model_id=AGENT_CONFIG["llm_model"],
                temperature=AGENT_CONFIG["temperature"],
                max_tokens=AGENT_CONFIG["max_tokens"]
            )
            
            # Initialize LLM manager and create LLM instance
            llm_manager = LLMProviderManager()
            await llm_manager.initialize()
            llm = await llm_manager.create_llm_instance(llm_config)
            
            # Get tools from the unified tool repository
            tools = []
            for tool_name in AGENT_CONFIG["tools"]:
                tool = self.orchestrator.tool_repository.get_tool(tool_name)
                if tool:
                    tools.append(tool)
                    print(f"   ✅ Tool loaded: {tool_name}")
                else:
                    print(f"   ⚠️ Tool not found: {tool_name}")
            
            print(f"   🛠️ {len(tools)} tools loaded successfully")
            
            # Create autonomous agent
            self.agent = create_autonomous_agent(
                name=AGENT_CONFIG["name"],
                description=AGENT_CONFIG["description"],
                llm=llm,
                tools=tools,
                autonomy_level=AutonomyLevel.PROACTIVE,
                learning_mode=LearningMode.ACTIVE,
                agent_id=self.agent_id,
                system_prompt=SYSTEM_PROMPT,
                memory_system=self.orchestrator.memory_system,
                rag_system=self.orchestrator.unified_rag
            )
           
            
            print("✅ Excel Analysis Agent created successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Agent creation failed: {str(e)}")
            logger.error("Agent creation failed", error=str(e))
            return False
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Parse Excel file and extract structured data."""
        try:
            print(f"📊 Parsing Excel file: {file_path}")
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            excel_data = {
                "file_path": file_path,
                "sheets": {},
                "metadata": {
                    "sheet_count": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                    "parsed_at": datetime.now().isoformat()
                }
            }
            
            # Parse each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = self._parse_sheet(sheet)
                excel_data["sheets"][sheet_name] = sheet_data
                print(f"   ✅ Parsed sheet: {sheet_name} ({sheet_data['row_count']} rows, {sheet_data['col_count']} cols)")
            
            workbook.close()
            print(f"✅ Excel file parsed successfully: {len(excel_data['sheets'])} sheets")
            return excel_data
            
        except Exception as e:
            print(f"❌ Excel parsing failed: {str(e)}")
            logger.error("Excel parsing failed", error=str(e))
            raise
    
    def _parse_sheet(self, sheet) -> Dict[str, Any]:
        """Parse individual Excel sheet."""
        data = []
        headers = []
        
        # Get sheet dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Extract data
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # First row as headers
                headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(row, 1)]
            else:
                # Data rows
                row_data = [cell if cell is not None else "" for cell in row]
                if any(str(cell).strip() for cell in row_data):  # Skip empty rows
                    data.append(row_data)
        
        return {
            "headers": headers,
            "data": data,
            "row_count": len(data),
            "col_count": len(headers),
            "max_row": max_row,
            "max_col": max_col
        }
    
    def extract_financial_metrics(self, excel_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key financial metrics from parsed Excel data."""
        try:
            print("🔍 Extracting financial metrics...")
            
            metrics = {
                "revenue_data": {},
                "cost_data": {},
                "profitability_data": {},
                "growth_metrics": {},
                "key_ratios": {}
            }
            
            # Analyze each sheet for financial data
            for sheet_name, sheet_data in excel_data["sheets"].items():
                sheet_metrics = self._analyze_sheet_for_metrics(sheet_name, sheet_data)
                
                # Categorize metrics by sheet purpose
                if "revenue" in sheet_name.lower() or "sales" in sheet_name.lower():
                    metrics["revenue_data"][sheet_name] = sheet_metrics
                elif "cost" in sheet_name.lower() or "expense" in sheet_name.lower():
                    metrics["cost_data"][sheet_name] = sheet_metrics
                elif "profit" in sheet_name.lower() or "roi" in sheet_name.lower():
                    metrics["profitability_data"][sheet_name] = sheet_metrics
                else:
                    # General analysis
                    metrics["key_ratios"][sheet_name] = sheet_metrics
            
            print("✅ Financial metrics extracted successfully")
            return metrics

        except Exception as e:
            print(f"❌ Financial metrics extraction failed: {str(e)}")
            logger.error("Financial metrics extraction failed", error=str(e))
            return {
                "revenue_data": {},
                "cost_data": {},
                "profitability_data": {},
                "growth_metrics": {},
                "key_ratios": {}
            }

    async def analyze_with_llm(self, excel_data: Dict[str, Any], financial_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Use LLM to perform intelligent analysis of Excel data."""
        try:
            print("🧠 Performing LLM-powered analysis...")

            # Prepare analysis prompt
            analysis_prompt = self._create_analysis_prompt(excel_data, financial_metrics)

            # Use the agent's LLM for analysis
            if self.agent and hasattr(self.agent, 'llm'):
                response = await self.agent.llm.ainvoke(analysis_prompt)
                analysis_text = response.content if hasattr(response, 'content') else str(response)
            else:
                # Fallback to business intelligence tool
                bi_tool = self.orchestrator.tool_repository.get_tool("business_intelligence")
                if bi_tool:
                    analysis_result = await bi_tool._arun(
                        analysis_type="financial",
                        business_context={
                            "excel_data": excel_data,
                            "financial_metrics": financial_metrics
                        },
                        time_horizon="1_year",
                        include_recommendations=True,
                        detail_level="comprehensive"
                    )
                    analysis_text = json.dumps(analysis_result, indent=2)
                else:
                    analysis_text = "LLM analysis not available"

            # Parse and structure the analysis
            structured_analysis = self._structure_analysis_response(analysis_text, financial_metrics)

            print("✅ LLM analysis completed successfully")
            return structured_analysis

        except Exception as e:
            print(f"❌ LLM analysis failed: {str(e)}")
            logger.error("LLM analysis failed", error=str(e))
            return {"error": str(e), "analysis": "Analysis failed"}

    def _create_analysis_prompt(self, excel_data: Dict[str, Any], financial_metrics: Dict[str, Any]) -> str:
        """Create comprehensive analysis prompt for LLM."""

        # Summarize Excel structure
        excel_summary = f"""
EXCEL FILE ANALYSIS REQUEST

File Structure:
- Total Sheets: {excel_data['metadata']['sheet_count']}
- Sheet Names: {', '.join(excel_data['metadata']['sheet_names'])}

Sheet Details:
"""

        for sheet_name, sheet_data in excel_data["sheets"].items():
            excel_summary += f"""
{sheet_name}:
- Rows: {sheet_data['row_count']}, Columns: {sheet_data['col_count']}
- Headers: {', '.join(sheet_data['headers'][:10])}{'...' if len(sheet_data['headers']) > 10 else ''}
"""

        # Add financial metrics summary
        metrics_summary = f"""
Financial Metrics Extracted:
- Revenue Data Sheets: {len(financial_metrics.get('revenue_data', {}))}
- Cost Data Sheets: {len(financial_metrics.get('cost_data', {}))}
- Profitability Sheets: {len(financial_metrics.get('profitability_data', {}))}
- Key Ratios Sheets: {len(financial_metrics.get('key_ratios', {}))}
"""

        prompt = f"""{excel_summary}

{metrics_summary}

ANALYSIS REQUEST:
Please provide a comprehensive business analysis of this Excel spreadsheet data. Focus on:

1. FINANCIAL HEALTH ASSESSMENT:
   - Overall revenue trends and patterns
   - Cost structure analysis
   - Profitability assessment
   - Cash flow implications

2. BUSINESS INSIGHTS:
   - Key performance indicators
   - Growth opportunities identified
   - Risk factors and concerns
   - Seasonal or cyclical patterns

3. TREND ANALYSIS:
   - Revenue growth trends
   - Cost optimization opportunities
   - Margin improvement potential
   - Performance benchmarks

4. ACTIONABLE RECOMMENDATIONS:
   - Strategic recommendations for growth
   - Cost optimization suggestions
   - Risk mitigation strategies
   - Investment priorities

5. ROI ANALYSIS:
   - Return on investment assessment
   - Profitability ratios analysis
   - Efficiency metrics evaluation
   - Performance vs industry standards

Please provide specific, data-driven insights with clear reasoning and actionable recommendations.
Format your response as a structured business analysis report.
"""

        return prompt

    def _structure_analysis_response(self, analysis_text: str, financial_metrics: Dict[str, Any]) -> Dict[str, Any]:
        """Structure the LLM analysis response into organized insights."""

        return {
            "analysis_summary": {
                "overall_assessment": self._extract_section(analysis_text, "FINANCIAL HEALTH ASSESSMENT"),
                "key_insights": self._extract_section(analysis_text, "BUSINESS INSIGHTS"),
                "trend_analysis": self._extract_section(analysis_text, "TREND ANALYSIS"),
                "recommendations": self._extract_section(analysis_text, "ACTIONABLE RECOMMENDATIONS"),
                "roi_analysis": self._extract_section(analysis_text, "ROI ANALYSIS")
            },
            "financial_metrics": financial_metrics,
            "analysis_metadata": {
                "analysis_date": datetime.now().isoformat(),
                "analysis_type": "comprehensive_excel_analysis",
                "confidence_level": "high"
            },
            "raw_analysis": analysis_text
        }

    def _extract_section(self, text: str, section_name: str) -> str:
        """Extract specific section from analysis text."""
        try:
            # Simple section extraction - in production, this could be more sophisticated
            lines = text.split('\n')
            section_lines = []
            in_section = False

            for line in lines:
                if section_name.upper() in line.upper():
                    in_section = True
                    continue
                elif in_section and any(keyword in line.upper() for keyword in ['ANALYSIS:', 'ASSESSMENT:', 'RECOMMENDATIONS:']):
                    if section_name.upper() not in line.upper():
                        break
                elif in_section:
                    section_lines.append(line.strip())

            return '\n'.join(section_lines).strip() if section_lines else f"Analysis for {section_name} not found in response."

        except Exception:
            return f"Could not extract {section_name} section."

    async def run_excel_analysis(self, excel_file_path: str) -> Dict[str, Any]:
        """Run complete Excel analysis workflow."""
        try:
            print(f"🚀 Starting Excel Analysis for: {excel_file_path}")

            # Step 1: Parse Excel file
            excel_data = self.parse_excel_file(excel_file_path)

            # Step 2: Extract financial metrics
            financial_metrics = self.extract_financial_metrics(excel_data)

            # Step 3: Store in RAG system
            if self.orchestrator.unified_rag:
                try:
                    await self.orchestrator.unified_rag.add_document(
                        content=json.dumps({
                            "excel_data": excel_data,
                            "financial_metrics": financial_metrics
                        }, indent=2),
                        metadata={
                            "type": "excel_analysis",
                            "file_path": excel_file_path,
                            "analyzed_at": datetime.now().isoformat()
                        },
                        collection=AGENT_CONFIG["collection_name"]
                    )
                    print("   📚 Data stored in RAG system")
                except Exception as e:
                    print(f"   ⚠️ RAG storage failed: {str(e)}")

            # Step 4: LLM-powered analysis
            llm_analysis = await self.analyze_with_llm(excel_data, financial_metrics)

            # Step 5: Compile final results
            result = {
                "status": "success",
                "file_path": excel_file_path,
                "excel_structure": excel_data["metadata"],
                "financial_metrics": financial_metrics,
                "llm_analysis": llm_analysis,
                "analysis_timestamp": datetime.now().isoformat()
            }

            print("✅ Excel Analysis completed successfully!")
            return result

        except Exception as e:
            print(f"❌ Excel analysis failed: {str(e)}")
            logger.error("Excel analysis failed", error=str(e))
            return {
                "status": "error",
                "error": str(e),
                "file_path": excel_file_path,
                "timestamp": datetime.now().isoformat()
            }

# ================================
# 🚀 MAIN LAUNCHER
# ================================

async def main():
    """Main launcher for the Excel Analysis Agent."""
    print("📊 EXCEL ANALYSIS AGENT")
    print("=" * 50)

    # Create and initialize agent
    agent_launcher = ExcelAnalysisAgent()

    # Initialize infrastructure
    if not await agent_launcher.initialize_infrastructure():
        print("❌ Failed to initialize infrastructure")
        return

    # Create agent
    if not await agent_launcher.create_agent():
        print("❌ Failed to create agent")
        return

    # Look for Excel files to analyze
    excel_files = list(Path("data/outputs").glob("*.xlsx"))

    if not excel_files:
        print("⚠️ No Excel files found in data/outputs/")
        print("💡 Run the Business Revenue Metrics Agent first to generate Excel files")
        return

    # Analyze each Excel file
    for excel_file in excel_files:
        print(f"\n📊 Analyzing: {excel_file.name}")
        result = await agent_launcher.run_excel_analysis(str(excel_file))

        if result["status"] == "success":
            print(f"✅ Analysis completed for {excel_file.name}")

            # Save analysis results
            analysis_file = excel_file.parent / f"{excel_file.stem}_analysis.json"
            with open(analysis_file, 'w') as f:
                json.dump(result, f, indent=2)
            print(f"   💾 Analysis saved to: {analysis_file}")
        else:
            print(f"❌ Analysis failed for {excel_file.name}: {result.get('error', 'Unknown error')}")

    print("\n🎉 Excel analysis completed!")

if __name__ == "__main__":
    asyncio.run(main())
    
    def _analyze_sheet_for_metrics(self, sheet_name: str, sheet_data: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze individual sheet for financial metrics."""
        metrics = {
            "sheet_name": sheet_name,
            "numeric_columns": [],
            "totals": {},
            "trends": {},
            "key_values": {}
        }
        
        headers = sheet_data["headers"]
        data = sheet_data["data"]
        
        # Identify numeric columns
        for col_idx, header in enumerate(headers):
            numeric_values = []
            for row in data:
                if col_idx < len(row):
                    try:
                        # Try to convert to float
                        value = str(row[col_idx]).replace('$', '').replace(',', '').replace('%', '')
                        if value and value != '':
                            numeric_values.append(float(value))
                    except (ValueError, TypeError):
                        continue
            
            if len(numeric_values) > 0:
                metrics["numeric_columns"].append(header)
                metrics["totals"][header] = sum(numeric_values)
                metrics["key_values"][header] = {
                    "sum": sum(numeric_values),
                    "average": sum(numeric_values) / len(numeric_values),
                    "min": min(numeric_values),
                    "max": max(numeric_values),
                    "count": len(numeric_values)
                }
        
        return metrics
