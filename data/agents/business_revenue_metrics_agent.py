#!/usr/bin/env python3
"""
🤖 TRUE AGENTIC BUSINESS REVENUE METRICS AGENT
==============================================
Autonomous agent that uses LLM reasoning and dynamic tool selection to generate
comprehensive business revenue metrics and Excel spreadsheets.

TRUE AGENTIC BEHAVIOR:
🧠 Uses LLM reasoning to decide how to approach tasks
🛠️ Dynamically selects tools from UnifiedToolRepository based on needs
🎯 Makes autonomous decisions about data generation strategies
📊 Adapts approach based on business type and requirements
🔄 Uses tool calling with autonomous planning and execution

ARCHITECTURE:
- AutonomousAgent with PROACTIVE autonomy level
- Dynamic tool selection from UnifiedToolRepository
- LLM-driven task planning and execution
- Tool-based Excel generation and analysis
- Autonomous decision making throughout the process
"""

import os
import sys
import asyncio
import json
import uuid
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from pathlib import Path

# Add the project root to the Python path
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

# Agent Configuration - TRUE AGENTIC SETUP
AGENT_CONFIG = {
    "name": "Agentic Business Revenue Metrics Agent",
    "description": "Autonomous agent that uses LLM reasoning and dynamic tool selection to generate comprehensive business revenue metrics and Excel spreadsheets",
    "agent_type": "autonomous_business_analysis",
    "llm_provider": "ollama",
    "llm_model": "llama3.2:latest",
    "temperature": 0.3,
    "max_tokens": 8000,
    "autonomy_level": "PROACTIVE",
    "learning_mode": "ACTIVE",
    "use_cases": [
        "business_analysis",
        "document_generation",
        "excel_processing",
        "financial_analysis",
        "data_generation"
    ],
    "capabilities": [
        "autonomous_reasoning",
        "tool_selection",
        "financial_analysis",
        "excel_generation",
        "business_intelligence"
    ],
    "memory_type": "UNIFIED",
    "rag_enabled": True,
    "collection_name": "agentic_business_revenue_metrics"
}

# Import the complete production infrastructure
from app.core.unified_system_orchestrator import get_enhanced_system_orchestrator
from app.agents.factory import AgentBuilderFactory, AgentBuilderConfig, AgentType, MemoryType
from app.agents.base.agent import AgentCapability
from app.agents.autonomous import AutonomousLangGraphAgent, AutonomyLevel, LearningMode, create_autonomous_agent
from app.llm.models import LLMConfig, ProviderType
from app.llm.manager import LLMProviderManager
from app.memory.unified_memory_system import UnifiedMemorySystem
from app.rag.core.unified_rag_system import UnifiedRAGSystem
from app.tools.unified_tool_repository import UnifiedToolRepository

import structlog
logger = structlog.get_logger(__name__)

# System prompt for TRUE AGENTIC behavior
SYSTEM_PROMPT = """You are an AUTONOMOUS Business Revenue Metrics Agent with advanced reasoning capabilities.

🎯 YOUR MISSION: Generate comprehensive business revenue metrics and Excel spreadsheets using autonomous decision-making and tool selection.

🧠 AUTONOMOUS CAPABILITIES:
- Use LLM reasoning to analyze business requirements
- Dynamically select appropriate tools from your tool repository
- Make intelligent decisions about data generation strategies
- Adapt your approach based on business type and context
- Execute multi-step workflows autonomously

🛠️ AVAILABLE TOOLS (use get_tools_for_use_case to access):
- business_intelligence: For financial analysis and business insights
- revolutionary_document_intelligence: For Excel generation and document processing
- revolutionary_web_scraper: For market research and data gathering

📊 AUTONOMOUS WORKFLOW:
1. ANALYZE the business requirements using LLM reasoning
2. SELECT appropriate tools based on the specific needs
3. GENERATE realistic business data using business intelligence tools
4. CREATE comprehensive Excel spreadsheets using document tools
5. PROVIDE intelligent analysis and insights

🎯 KEY BEHAVIORS:
- Always reason through your approach before taking action
- Use tools dynamically based on the specific requirements
- Make autonomous decisions about data generation strategies
- Provide comprehensive analysis with your outputs
- Adapt your methodology based on business type and size

Remember: You are NOT following a script - you are making intelligent, autonomous decisions at each step!"""

class AgenticBusinessRevenueMetricsAgent:
    """Production-ready business revenue metrics agent."""
    
    def __init__(self):
        self.orchestrator = None
        self.agent = None
        self.agent_id = f"business_revenue_metrics_{uuid.uuid4().hex[:8]}"
        self.running = False
        
    async def initialize_infrastructure(self):
        """Initialize the complete production infrastructure."""
        try:
            print("🔧 Initializing Business Revenue Metrics Infrastructure...")
            
            # Get the enhanced system orchestrator
            self.orchestrator = get_enhanced_system_orchestrator()
            await self.orchestrator.initialize()
            
            print("✅ Infrastructure initialized successfully!")
            print(f"   🧠 Memory System: {type(self.orchestrator.memory_system).__name__}")
            print(f"   📚 RAG System: {type(self.orchestrator.unified_rag).__name__}")
            print(f"   🛠️ Tool Repository: {self.orchestrator.tool_repository.stats['total_tools']} tools available")
            
            return True
            
        except Exception as e:
            print(f"❌ Infrastructure initialization failed: {str(e)}")
            logger.error("Infrastructure initialization failed", error=str(e))
            return False
    
    async def create_agent(self):
        """Create the business revenue metrics agent."""
        try:
            print(f"🤖 Creating {AGENT_CONFIG['name']}...")
            
            # Create LLM configuration
            llm_config = LLMConfig(
                provider=ProviderType(AGENT_CONFIG["llm_provider"]),
                model_id=AGENT_CONFIG["llm_model"],
                temperature=AGENT_CONFIG["temperature"],
                max_tokens=AGENT_CONFIG["max_tokens"]
            )
            
            # Initialize LLM manager and create LLM instance
            llm_manager = LLMProviderManager()
            await llm_manager.initialize()
            llm = await llm_manager.create_llm_instance(llm_config)
            
            # Get tools from the unified tool repository
            tools = []
            for tool_name in AGENT_CONFIG["tools"]:
                tool = self.orchestrator.tool_repository.get_tool(tool_name)
                if tool:
                    tools.append(tool)
                    print(f"   ✅ Tool loaded: {tool_name}")
                else:
                    print(f"   ⚠️ Tool not found: {tool_name}")
            
            print(f"   🛠️ {len(tools)} tools loaded successfully")
            
            # Create autonomous agent
            self.agent = create_autonomous_agent(
                name=AGENT_CONFIG["name"],
                description=AGENT_CONFIG["description"],
                llm=llm,
                tools=tools,
                autonomy_level=AutonomyLevel.PROACTIVE,
                learning_mode=LearningMode.ACTIVE,
                agent_id=self.agent_id,
                system_prompt=SYSTEM_PROMPT,
                memory_system=self.orchestrator.memory_system,
                rag_system=self.orchestrator.unified_rag
            )
            
            print("✅ Business Revenue Metrics Agent created successfully!")
            return True
            
        except Exception as e:
            print(f"❌ Agent creation failed: {str(e)}")
            logger.error("Agent creation failed", error=str(e))
            return False

    async def execute_agentic_task(self, business_type: str = "technology", company_size: str = "medium") -> Dict[str, Any]:
        """
        🤖 EXECUTE TASK WITH TRUE AGENTIC BEHAVIOR

        This method uses the autonomous agent to reason through the task and make
        intelligent decisions about tool selection and execution strategy.
        """
        try:
            if not self.agent:
                print("❌ Agent not initialized")
                return {"status": "error", "error": "Agent not initialized"}

            print(f"🤖 Starting AGENTIC execution for {business_type} business ({company_size} size)")

            # Create the task prompt that will trigger autonomous reasoning
            task_prompt = f"""
            🎯 AUTONOMOUS TASK: Generate comprehensive business revenue metrics and Excel spreadsheet

            📋 REQUIREMENTS:
            - Business Type: {business_type}
            - Company Size: {company_size}
            - Output: Professional Excel spreadsheet with multiple sheets
            - Analysis: Comprehensive financial analysis and insights

            🧠 YOUR AUTONOMOUS APPROACH:
            1. REASON about the specific requirements for this business type and size
            2. SELECT appropriate tools from your repository based on the needs
            3. GENERATE realistic business data using business intelligence tools
            4. CREATE comprehensive Excel spreadsheet using document tools
            5. PROVIDE intelligent analysis and insights

            🛠️ AVAILABLE TOOLS (use as needed):
            - business_intelligence: For financial analysis and business insights
            - revolutionary_document_intelligence: For Excel generation and document processing
            - revolutionary_web_scraper: For market research and data gathering

            Remember: You are making autonomous decisions at each step. Reason through your approach,
            select tools dynamically, and adapt your strategy based on the specific business context.

            Execute this task autonomously using your reasoning capabilities and tool selection!
            """

            print("🧠 Sending task to autonomous agent for reasoning and execution...")

            # Execute the task using the autonomous agent
            result = await self.agent.execute_task(
                task=task_prompt,
                session_id=f"business_revenue_{uuid.uuid4().hex[:8]}"
            )

            print("✅ Autonomous agent execution completed!")

            return {
                "status": "success",
                "agent_result": result,
                "business_type": business_type,
                "company_size": company_size,
                "execution_timestamp": datetime.now().isoformat()
            }

        except Exception as e:
            print(f"❌ Agentic execution failed: {str(e)}")
            logger.error("Agentic execution failed", error=str(e))
            return {
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }
    
    def generate_sample_business_data(self, business_type: str = "technology", company_size: str = "medium") -> Dict[str, Any]:
        """Generate realistic sample business data for demonstration."""
        
        # Base revenue multipliers by company size
        size_multipliers = {
            "small": 1.0,
            "medium": 5.0,
            "large": 25.0,
            "enterprise": 100.0
        }
        
        base_multiplier = size_multipliers.get(company_size, 5.0)
        
        # Generate quarterly revenue data
        quarterly_data = {}
        base_revenue = 1000000 * base_multiplier  # Base $1M for medium company
        
        for i, quarter in enumerate(REVENUE_METRICS_CONFIG["time_periods"]):
            # Add seasonal variation and growth
            seasonal_factor = [0.9, 1.1, 1.0, 1.2][i]  # Q4 typically higher
            growth_factor = 1 + (i * 0.05)  # 5% quarterly growth
            
            quarter_revenue = base_revenue * seasonal_factor * growth_factor
            
            quarterly_data[quarter] = {
                "total_revenue": quarter_revenue,
                "revenue_streams": {
                    stream: quarter_revenue * random.uniform(0.1, 0.4) 
                    for stream in REVENUE_METRICS_CONFIG["revenue_streams"]
                },
                "costs": {
                    cost: quarter_revenue * random.uniform(0.05, 0.25)
                    for cost in REVENUE_METRICS_CONFIG["cost_categories"]
                }
            }
        
        return {
            "company_info": {
                "name": f"Sample {business_type.title()} Company",
                "type": business_type,
                "size": company_size,
                "fiscal_year": datetime.now().year
            },
            "quarterly_data": quarterly_data,
            "generated_at": datetime.now().isoformat()
        }
    
    async def generate_revenue_metrics_excel(self, business_data: Dict[str, Any], output_path: str = None) -> str:
        """Generate comprehensive revenue metrics Excel spreadsheet."""
        try:
            print("📊 Generating Revenue Metrics Excel Spreadsheet...")

            # Generate Excel file directly using openpyxl
            excel_bytes = self._create_excel_workbook(business_data)

            # Save to file
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"data/outputs/business_revenue_metrics_{timestamp}.xlsx"

            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, 'wb') as f:
                f.write(excel_bytes)

            print(f"✅ Excel spreadsheet generated: {output_path}")
            return output_path

        except Exception as e:
            print(f"❌ Excel generation failed: {str(e)}")
            logger.error("Excel generation failed", error=str(e))
            raise

    def _create_excel_workbook(self, business_data: Dict[str, Any]) -> bytes:
        """Create Excel workbook with business revenue metrics."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO

            # Create workbook
            workbook = openpyxl.Workbook()

            # Remove default sheet
            workbook.remove(workbook.active)

            # Prepare data
            company_info = business_data["company_info"]
            quarterly_data = business_data["quarterly_data"]
            yearly_totals = self._calculate_yearly_totals(quarterly_data)
            roi_metrics = self._calculate_roi_metrics(yearly_totals)

            # Create sheets
            self._create_summary_sheet_excel(workbook, company_info, yearly_totals, roi_metrics)
            self._create_revenue_sheet_excel(workbook, quarterly_data)
            self._create_cost_sheet_excel(workbook, quarterly_data)
            self._create_profitability_sheet_excel(workbook, quarterly_data, yearly_totals)
            self._create_roi_sheet_excel(workbook, roi_metrics)

            # Save to bytes
            output = BytesIO()
            workbook.save(output)
            output.seek(0)

            return output.getvalue()

        except Exception as e:
            logger.error(f"Excel workbook creation failed: {str(e)}")
            raise
    
    def _prepare_excel_data(self, business_data: Dict[str, Any]) -> Dict[str, Any]:
        """Prepare data structure for Excel generation."""
        
        company_info = business_data["company_info"]
        quarterly_data = business_data["quarterly_data"]
        
        # Calculate yearly totals and metrics
        yearly_totals = self._calculate_yearly_totals(quarterly_data)
        roi_metrics = self._calculate_roi_metrics(yearly_totals)
        
        return {
            "sheets": [
                {
                    "name": "Executive Summary",
                    "data": self._create_summary_sheet(company_info, yearly_totals, roi_metrics)
                },
                {
                    "name": "Revenue Analysis", 
                    "data": self._create_revenue_sheet(quarterly_data)
                },
                {
                    "name": "Cost Analysis",
                    "data": self._create_cost_sheet(quarterly_data)
                },
                {
                    "name": "Profitability",
                    "data": self._create_profitability_sheet(quarterly_data, yearly_totals)
                },
                {
                    "name": "ROI Metrics",
                    "data": self._create_roi_sheet(roi_metrics)
                }
            ]
        }

    def _create_summary_sheet_excel(self, workbook, company_info, yearly_totals, roi_metrics):
        """Create executive summary sheet."""
        sheet = workbook.create_sheet("Executive Summary")

        # Header style
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, size=14)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Company info
        sheet['A1'] = f"{company_info['name']} - Annual Revenue Report"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet['A2'] = f"Business Type: {company_info['type'].title()}"
        sheet['A3'] = f"Company Size: {company_info['size'].title()}"
        sheet['A4'] = f"Report Year: {company_info['fiscal_year']}"

        # Key metrics
        sheet['A6'] = "KEY FINANCIAL METRICS"
        sheet['A6'].font = header_font
        sheet['A6'].fill = header_fill

        sheet['A7'] = "Total Revenue"
        sheet['B7'] = f"${yearly_totals['total_revenue']:,.2f}"
        sheet['A8'] = "Total Costs"
        sheet['B8'] = f"${yearly_totals['total_costs']:,.2f}"
        sheet['A9'] = "Net Profit"
        sheet['B9'] = f"${yearly_totals['net_profit']:,.2f}"
        sheet['A10'] = "Profit Margin"
        sheet['B10'] = f"{yearly_totals['profit_margin']:.2f}%"

        # ROI metrics
        sheet['A12'] = "ROI ANALYSIS"
        sheet['A12'].font = header_font
        sheet['A12'].fill = header_fill

        sheet['A13'] = "Return on Investment"
        sheet['B13'] = f"{roi_metrics['roi']:.2f}%"
        sheet['A14'] = "Return on Assets"
        sheet['B14'] = f"{roi_metrics['roa']:.2f}%"
        sheet['A15'] = "Return on Equity"
        sheet['B15'] = f"{roi_metrics['roe']:.2f}%"

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_revenue_sheet_excel(self, workbook, quarterly_data):
        """Create revenue analysis sheet."""
        sheet = workbook.create_sheet("Revenue Analysis")

        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        # Headers
        headers = ["Quarter", "Product Sales", "Service Revenue", "Other Revenue", "Total Revenue"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row, (quarter, quarter_data) in enumerate(quarterly_data.items(), 2):
            revenue_streams = quarter_data["revenue_streams"]
            sheet.cell(row=row, column=1, value=quarter)
            sheet.cell(row=row, column=2, value=revenue_streams.get("product_sales", 0))
            sheet.cell(row=row, column=3, value=revenue_streams.get("service_revenue", 0))
            sheet.cell(row=row, column=4, value=revenue_streams.get("other_revenue", 0))
            sheet.cell(row=row, column=5, value=quarter_data["total_revenue"])

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_cost_sheet_excel(self, workbook, quarterly_data):
        """Create cost analysis sheet."""
        sheet = workbook.create_sheet("Cost Analysis")

        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")

        # Headers
        headers = ["Quarter", "COGS", "Operating Expenses", "Marketing", "R&D", "Total Costs"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row, (quarter, quarter_data) in enumerate(quarterly_data.items(), 2):
            costs = quarter_data["costs"]
            total_costs = sum(costs.values())
            sheet.cell(row=row, column=1, value=quarter)
            sheet.cell(row=row, column=2, value=costs.get("cogs", 0))
            sheet.cell(row=row, column=3, value=costs.get("operating_expenses", 0))
            sheet.cell(row=row, column=4, value=costs.get("marketing", 0))
            sheet.cell(row=row, column=5, value=costs.get("rd", 0))
            sheet.cell(row=row, column=6, value=total_costs)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_profitability_sheet_excel(self, workbook, quarterly_data, yearly_totals):
        """Create profitability analysis sheet."""
        sheet = workbook.create_sheet("Profitability")

        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")

        # Headers
        headers = ["Quarter", "Revenue", "Costs", "Gross Profit", "Profit Margin %"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data
        for row, (quarter, quarter_data) in enumerate(quarterly_data.items(), 2):
            revenue = quarter_data["total_revenue"]
            costs = sum(quarter_data["costs"].values())
            profit = revenue - costs
            margin = (profit / revenue * 100) if revenue > 0 else 0

            sheet.cell(row=row, column=1, value=quarter)
            sheet.cell(row=row, column=2, value=revenue)
            sheet.cell(row=row, column=3, value=costs)
            sheet.cell(row=row, column=4, value=profit)
            sheet.cell(row=row, column=5, value=f"{margin:.2f}%")

        # Yearly totals
        row = len(quarterly_data) + 3
        sheet.cell(row=row, column=1, value="YEARLY TOTAL").font = Font(bold=True)
        sheet.cell(row=row, column=2, value=yearly_totals["total_revenue"])
        sheet.cell(row=row, column=3, value=yearly_totals["total_costs"])
        sheet.cell(row=row, column=4, value=yearly_totals["net_profit"])
        sheet.cell(row=row, column=5, value=f"{yearly_totals['profit_margin']:.2f}%")

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _create_roi_sheet_excel(self, workbook, roi_metrics):
        """Create ROI analysis sheet."""
        sheet = workbook.create_sheet("ROI Metrics")

        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

        # Headers
        sheet['A1'] = "ROI ANALYSIS"
        sheet['A1'].font = Font(bold=True, size=16)

        # Metrics
        metrics = [
            ("Return on Investment (ROI)", f"{roi_metrics['roi']:.2f}%"),
            ("Return on Assets (ROA)", f"{roi_metrics['roa']:.2f}%"),
            ("Return on Equity (ROE)", f"{roi_metrics['roe']:.2f}%"),
            ("Gross Profit Margin", f"{roi_metrics['gross_margin']:.2f}%"),
            ("Net Profit Margin", f"{roi_metrics['net_margin']:.2f}%"),
            ("Asset Turnover", f"{roi_metrics['asset_turnover']:.2f}"),
            ("Equity Multiplier", f"{roi_metrics['equity_multiplier']:.2f}")
        ]

        for row, (metric, value) in enumerate(metrics, 3):
            sheet.cell(row=row, column=1, value=metric).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=value)

        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _calculate_yearly_totals(self, quarterly_data: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate yearly totals from quarterly data."""
        yearly_totals = {
            "total_revenue": 0,
            "total_costs": 0,
            "revenue_by_stream": {},
            "costs_by_category": {}
        }

        # Sum up quarterly data
        for quarter, data in quarterly_data.items():
            yearly_totals["total_revenue"] += data["total_revenue"]

            # Sum revenue streams
            for stream, amount in data["revenue_streams"].items():
                if stream not in yearly_totals["revenue_by_stream"]:
                    yearly_totals["revenue_by_stream"][stream] = 0
                yearly_totals["revenue_by_stream"][stream] += amount

            # Sum costs
            quarter_costs = sum(data["costs"].values())
            yearly_totals["total_costs"] += quarter_costs

            for category, amount in data["costs"].items():
                if category not in yearly_totals["costs_by_category"]:
                    yearly_totals["costs_by_category"][category] = 0
                yearly_totals["costs_by_category"][category] += amount

        yearly_totals["net_profit"] = yearly_totals["total_revenue"] - yearly_totals["total_costs"]
        yearly_totals["profit_margin"] = (yearly_totals["net_profit"] / yearly_totals["total_revenue"]) * 100 if yearly_totals["total_revenue"] > 0 else 0

        return yearly_totals

    def _calculate_roi_metrics(self, yearly_totals: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate ROI and financial metrics."""
        revenue = yearly_totals["total_revenue"]
        costs = yearly_totals["total_costs"]
        profit = yearly_totals["net_profit"]

        # Assume some assets and equity for ROI calculations
        assumed_assets = revenue * 0.8  # 80% of revenue as assets
        assumed_equity = revenue * 0.4  # 40% of revenue as equity

        return {
            "roi": (profit / costs) * 100 if costs > 0 else 0,
            "roa": (profit / assumed_assets) * 100 if assumed_assets > 0 else 0,
            "roe": (profit / assumed_equity) * 100 if assumed_equity > 0 else 0,
            "gross_margin": yearly_totals["profit_margin"],
            "net_margin": yearly_totals["profit_margin"],
            "asset_turnover": revenue / assumed_assets if assumed_assets > 0 else 0,
            "equity_multiplier": assumed_assets / assumed_equity if assumed_equity > 0 else 0
        }

    def _create_summary_sheet(self, company_info: Dict[str, Any], yearly_totals: Dict[str, Any], roi_metrics: Dict[str, Any]) -> List[List[Any]]:
        """Create executive summary sheet data."""
        return [
            ["EXECUTIVE SUMMARY", "", "", ""],
            ["", "", "", ""],
            ["Company Information", "", "", ""],
            ["Company Name", company_info["name"], "", ""],
            ["Business Type", company_info["type"], "", ""],
            ["Company Size", company_info["size"], "", ""],
            ["Fiscal Year", company_info["fiscal_year"], "", ""],
            ["", "", "", ""],
            ["Financial Performance", "", "", ""],
            ["Total Revenue", f"${yearly_totals['total_revenue']:,.2f}", "", ""],
            ["Total Costs", f"${yearly_totals['total_costs']:,.2f}", "", ""],
            ["Gross Profit", f"${yearly_totals['gross_profit']:,.2f}", "", ""],
            ["Profit Margin", f"{yearly_totals['profit_margin']:.2f}%", "", ""],
            ["", "", "", ""],
            ["Key ROI Metrics", "", "", ""],
            ["Return on Investment", f"{roi_metrics['roi_percent']:.2f}%", "", ""],
            ["Return on Assets", f"{roi_metrics['roa_percent']:.2f}%", "", ""],
            ["Return on Equity", f"{roi_metrics['roe_percent']:.2f}%", "", ""],
            ["Revenue Growth Rate", f"{roi_metrics['revenue_growth_rate']:.2f}%", "", ""],
            ["Cost Efficiency Ratio", f"{roi_metrics['cost_efficiency_ratio']:.2f}%", "", ""]
        ]

    def _create_revenue_sheet(self, quarterly_data: Dict[str, Any]) -> List[List[Any]]:
        """Create revenue analysis sheet data."""
        headers = ["Revenue Stream"] + list(quarterly_data.keys()) + ["Total"]
        data = [headers]

        # Get all revenue streams
        all_streams = set()
        for quarter_data in quarterly_data.values():
            all_streams.update(quarter_data["revenue_streams"].keys())

        # Create rows for each revenue stream
        for stream in sorted(all_streams):
            row = [stream]
            total = 0
            for quarter in quarterly_data.keys():
                amount = quarterly_data[quarter]["revenue_streams"].get(stream, 0)
                row.append(f"${amount:,.2f}")
                total += amount
            row.append(f"${total:,.2f}")
            data.append(row)

        # Add total row
        total_row = ["TOTAL"]
        grand_total = 0
        for quarter in quarterly_data.keys():
            quarter_total = sum(quarterly_data[quarter]["revenue_streams"].values())
            total_row.append(f"${quarter_total:,.2f}")
            grand_total += quarter_total
        total_row.append(f"${grand_total:,.2f}")
        data.append(total_row)

        return data

    def _create_cost_sheet(self, quarterly_data: Dict[str, Any]) -> List[List[Any]]:
        """Create cost analysis sheet data."""
        headers = ["Cost Category"] + list(quarterly_data.keys()) + ["Total"]
        data = [headers]

        # Get all cost categories
        all_categories = set()
        for quarter_data in quarterly_data.values():
            all_categories.update(quarter_data["costs"].keys())

        # Create rows for each cost category
        for category in sorted(all_categories):
            row = [category]
            total = 0
            for quarter in quarterly_data.keys():
                amount = quarterly_data[quarter]["costs"].get(category, 0)
                row.append(f"${amount:,.2f}")
                total += amount
            row.append(f"${total:,.2f}")
            data.append(row)

        # Add total row
        total_row = ["TOTAL"]
        grand_total = 0
        for quarter in quarterly_data.keys():
            quarter_total = sum(quarterly_data[quarter]["costs"].values())
            total_row.append(f"${quarter_total:,.2f}")
            grand_total += quarter_total
        total_row.append(f"${grand_total:,.2f}")
        data.append(total_row)

        return data

    def _create_profitability_sheet(self, quarterly_data: Dict[str, Any], yearly_totals: Dict[str, Any]) -> List[List[Any]]:
        """Create profitability analysis sheet data."""
        headers = ["Metric"] + list(quarterly_data.keys()) + ["Total"]
        data = [headers]

        # Revenue row
        revenue_row = ["Revenue"]
        for quarter in quarterly_data.keys():
            revenue_row.append(f"${quarterly_data[quarter]['total_revenue']:,.2f}")
        revenue_row.append(f"${yearly_totals['total_revenue']:,.2f}")
        data.append(revenue_row)

        # Costs row
        costs_row = ["Total Costs"]
        for quarter in quarterly_data.keys():
            quarter_costs = sum(quarterly_data[quarter]["costs"].values())
            costs_row.append(f"${quarter_costs:,.2f}")
        costs_row.append(f"${yearly_totals['total_costs']:,.2f}")
        data.append(costs_row)

        # Profit row
        profit_row = ["Gross Profit"]
        for quarter in quarterly_data.keys():
            quarter_profit = quarterly_data[quarter]['total_revenue'] - sum(quarterly_data[quarter]["costs"].values())
            profit_row.append(f"${quarter_profit:,.2f}")
        profit_row.append(f"${yearly_totals['gross_profit']:,.2f}")
        data.append(profit_row)

        # Margin row
        margin_row = ["Profit Margin %"]
        for quarter in quarterly_data.keys():
            quarter_revenue = quarterly_data[quarter]['total_revenue']
            quarter_costs = sum(quarterly_data[quarter]["costs"].values())
            quarter_profit = quarter_revenue - quarter_costs
            margin = (quarter_profit / quarter_revenue) * 100 if quarter_revenue > 0 else 0
            margin_row.append(f"{margin:.2f}%")
        margin_row.append(f"{yearly_totals['profit_margin']:.2f}%")
        data.append(margin_row)

        return data

    def _create_roi_sheet(self, roi_metrics: Dict[str, Any]) -> List[List[Any]]:
        """Create ROI metrics sheet data."""
        return [
            ["ROI METRICS ANALYSIS", "", "", ""],
            ["", "", "", ""],
            ["Metric", "Value", "Description", ""],
            ["Return on Investment (ROI)", f"{roi_metrics['roi_percent']:.2f}%", "Profit relative to costs", ""],
            ["Return on Assets (ROA)", f"{roi_metrics['roa_percent']:.2f}%", "Profit relative to assets", ""],
            ["Return on Equity (ROE)", f"{roi_metrics['roe_percent']:.2f}%", "Profit relative to equity", ""],
            ["Gross Margin", f"{roi_metrics['gross_margin_percent']:.2f}%", "Profit margin percentage", ""],
            ["Revenue Growth Rate", f"{roi_metrics['revenue_growth_rate']:.2f}%", "Year-over-year growth", ""],
            ["Cost Efficiency Ratio", f"{roi_metrics['cost_efficiency_ratio']:.2f}%", "Costs as % of revenue", ""],
            ["", "", "", ""],
            ["PERFORMANCE INDICATORS", "", "", ""],
            ["ROI Rating", "Excellent" if roi_metrics['roi_percent'] > 20 else "Good" if roi_metrics['roi_percent'] > 10 else "Fair", "", ""],
            ["Profitability", "High" if roi_metrics['gross_margin_percent'] > 25 else "Medium" if roi_metrics['gross_margin_percent'] > 15 else "Low", "", ""],
            ["Growth Outlook", "Strong" if roi_metrics['revenue_growth_rate'] > 15 else "Moderate" if roi_metrics['revenue_growth_rate'] > 5 else "Slow", "", ""]
        ]

    async def run_analysis(self, business_type: str = "technology", company_size: str = "medium", custom_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """Run complete business revenue metrics analysis."""
        try:
            print(f"🚀 Starting Business Revenue Metrics Analysis...")
            print(f"   📊 Business Type: {business_type}")
            print(f"   🏢 Company Size: {company_size}")

            # Generate or use provided business data
            if custom_data:
                business_data = custom_data
                print("   📋 Using provided business data")
            else:
                business_data = self.generate_sample_business_data(business_type, company_size)
                print("   🎲 Generated sample business data")

            # Generate Excel spreadsheet
            excel_path = await self.generate_revenue_metrics_excel(business_data)

            # Store in RAG system for future analysis
            if self.orchestrator.unified_rag:
                try:
                    await self.orchestrator.unified_rag.add_document(
                        content=json.dumps(business_data, indent=2),
                        metadata={
                            "type": "business_revenue_metrics",
                            "business_type": business_type,
                            "company_size": company_size,
                            "excel_path": excel_path,
                            "generated_at": datetime.now().isoformat()
                        },
                        collection=AGENT_CONFIG["collection_name"]
                    )
                    print("   📚 Data stored in RAG system")
                except Exception as e:
                    print(f"   ⚠️ RAG storage failed: {str(e)}")

            # Use business intelligence tool for additional analysis
            bi_tool = self.orchestrator.tool_repository.get_tool("business_intelligence")
            if bi_tool:
                try:
                    analysis_result = await bi_tool._arun(
                        analysis_type="financial",
                        business_context={
                            "business_data": business_data,
                            "excel_path": excel_path,
                            "analysis_timestamp": datetime.now().isoformat()
                        },
                        time_horizon="1_year",
                        include_recommendations=True,
                        detail_level="comprehensive"
                    )
                    print("   🧠 Business intelligence analysis completed")
                except Exception as e:
                    print(f"   ⚠️ Business intelligence analysis failed: {str(e)}")
                    analysis_result = None
            else:
                analysis_result = None

            result = {
                "status": "success",
                "excel_path": excel_path,
                "business_data": business_data,
                "analysis_result": analysis_result,
                "generated_at": datetime.now().isoformat()
            }

            print("✅ Business Revenue Metrics Analysis completed successfully!")
            print(f"   📄 Excel file: {excel_path}")

            return result

        except Exception as e:
            print(f"❌ Analysis failed: {str(e)}")
            logger.error("Business revenue metrics analysis failed", error=str(e))
            return {
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }

# ================================
# 🚀 MAIN LAUNCHER
# ================================

async def main():
    """Main launcher for the Business Revenue Metrics Agent."""
    print("🏢 BUSINESS REVENUE METRICS AGENT")
    print("=" * 50)

    # Create and initialize agent
    agent_launcher = BusinessRevenueMetricsAgent()

    # Initialize infrastructure
    if not await agent_launcher.initialize_infrastructure():
        print("❌ Failed to initialize infrastructure")
        return

    # Create agent
    if not await agent_launcher.create_agent():
        print("❌ Failed to create agent")
        return

    # Run analysis with different scenarios
    scenarios = [
        {"business_type": "technology", "company_size": "medium"},
        {"business_type": "retail", "company_size": "large"},
        {"business_type": "manufacturing", "company_size": "small"}
    ]

    for i, scenario in enumerate(scenarios, 1):
        print(f"\n📊 Running Scenario {i}: {scenario['business_type'].title()} - {scenario['company_size'].title()}")
        result = await agent_launcher.run_analysis(**scenario)

        if result["status"] == "success":
            print(f"✅ Scenario {i} completed successfully")
        else:
            print(f"❌ Scenario {i} failed: {result.get('error', 'Unknown error')}")

        # Small delay between scenarios
        await asyncio.sleep(2)

    print("\n🎉 All scenarios completed!")
    print("📁 Check the data/outputs/ directory for generated Excel files")

if __name__ == "__main__":
    asyncio.run(main())
